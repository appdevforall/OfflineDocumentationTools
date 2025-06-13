import sqlite3
import openpyxl
import json

class TooltipDatabase:
    SCHEMA = """
        CREATE TABLE IF NOT EXISTS `ide_tooltip_table` (
        `tooltipCategory` TEXT NOT NULL, 
        `tooltipTag` TEXT NOT NULL, 
        `tooltipSummary` TEXT NOT NULL, 
        `tooltipDetail` TEXT NOT NULL, 
        `tooltipButtons` TEXT NOT NULL, 
        PRIMARY KEY(`tooltipCategory`, `tooltipTag`));
    """
    
    REQUIRED_HEADERS = [
        'tag', 'category', 'summary', 'detail',
        'buttonDescr1', 'buttonURI1',
        'buttonDescr2', 'buttonURI2',
        'buttonDescr3', 'buttonURI3'
    ]
    
    REQUIRED_COLUMNS = ['tag', 'category', 'summary', 'buttonDescr1', 'buttonURI1']
    
    def __init__(self, db_path: str, xlsx_path: str):
        self.db_path = db_path
        self.xlsx_path = xlsx_path
        # Create the table if it doesn't exist
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute(self.SCHEMA)
        conn.commit()
        conn.close()

    def validate_row(self, row, headers, row_idx):
        for col_name in self.REQUIRED_COLUMNS:
            col_idx = headers.index(col_name)
            value = row[col_idx].value
            # Convert boolean False to string "False"
            if isinstance(value, bool):
                value = str(value)
            if not value:
                raise ValueError(f"Empty cell found in required column '{col_name}' at row {row_idx}")

    def _transform_kotlin_url(self, uri):
        if uri and uri.startswith("https://kotlinlang.org/docs"):
            return uri.replace("https://kotlinlang.org/docs", "http://localhost:6174/KotlinDocs/html")
        return uri

    def _create_button_hash(self, descr, uri):
        if descr and uri:
            return {
                "first": descr,
                "second": self._transform_kotlin_url(uri)
            }
        return None

    def read_xlsx(self):
        # Try to open the xlsx file using openpyxl
        wb = openpyxl.load_workbook(self.xlsx_path)
        ws = wb.active
        
        # Get the headers from the first row
        headers = [cell.value for cell in ws[1]]
        
        # Validate headers
        if headers != self.REQUIRED_HEADERS:
            raise ValueError(f"Invalid headers. Expected {self.REQUIRED_HEADERS}, got {headers}")
        
        # Connect to the database
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        skipped_rows = []
        
        try:
            # Check each row for empty cells in required columns and insert data
            for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):  # Start from row 2 (after headers)
                try:
                    self.validate_row(row, headers, row_idx)
                    
                    # Get values from the row
                    category = row[headers.index('category')].value
                    tag = row[headers.index('tag')].value
                    tag = ' ' if tag is None or tag == '' else str(tag)
                    summary = row[headers.index('summary')].value
                    detail = row[headers.index('detail')].value or ' '  # Convert empty detail to space
                    
                    # Create button hashes
                    button1_hash = self._create_button_hash(
                        row[headers.index('buttonDescr1')].value,
                        row[headers.index('buttonURI1')].value
                    )
                    
                    button2_hash = self._create_button_hash(
                        row[headers.index('buttonDescr2')].value,
                        row[headers.index('buttonURI2')].value
                    )
                    
                    button3_hash = self._create_button_hash(
                        row[headers.index('buttonDescr3')].value,
                        row[headers.index('buttonURI3')].value
                    )
                    
                    # Build button tuple
                    button_tuple = (button1_hash,)
                    if button2_hash:
                        button_tuple = button_tuple + (button2_hash,)
                    if button3_hash:
                        button_tuple = button_tuple + (button3_hash,)
                    
                    # Convert tuple to JSON string for storage
                    button_json = json.dumps(button_tuple)
                    
                    # Insert into database
                    cursor.execute("""
                        INSERT OR REPLACE INTO ide_tooltip_table 
                        (tooltipCategory, tooltipTag, tooltipSummary, tooltipDetail, tooltipButtons) 
                        VALUES (?, ?, ?, ?, ?)
                    """, (category, tag, summary, detail, button_json))
                    
                except ValueError as e:
                    # Log the skipped row
                    skipped_rows.append((row_idx, str(e)))
                    continue
            
            # Commit the changes
            conn.commit()
            
            # Return information about skipped rows
            return {
                'skipped_rows': skipped_rows,
                'total_rows': len(list(ws.iter_rows(min_row=2))),
                'processed_rows': len(list(ws.iter_rows(min_row=2))) - len(skipped_rows)
            }
            
        finally:
            # Always close the connection
            conn.close()
    
