import unittest
import sqlite3
import tempfile
import os
from tooltips import TooltipDatabase
from openpyxl import Workbook

class TestTooltipDatabase(unittest.TestCase):
    def test_create_database_with_paths(self):
        # Create a new TooltipDatabase instance with test paths
        db_path = "test.db"
        xlsx_path = "test.xlsx"
        db = TooltipDatabase(db_path, xlsx_path)
        
        # Verify the database object was created
        self.assertIsNotNone(db)
        self.assertIsInstance(db, TooltipDatabase)
        
        # Verify the paths were stored correctly
        self.assertEqual(db.db_path, db_path)
        self.assertEqual(db.xlsx_path, xlsx_path)

    def test_ide_tooltip_table_exists(self):
        # Use a temporary file for the database
        with tempfile.NamedTemporaryFile(delete=False) as tmp:
            db_path = tmp.name
        xlsx_path = "dummy.xlsx"
        try:
            db = TooltipDatabase(db_path, xlsx_path)
            # Connect to the database and check for the table
            conn = sqlite3.connect(db_path)
            cursor = conn.cursor()
            cursor.execute("""
                SELECT name FROM sqlite_master WHERE type='table' AND name='ide_tooltip_table';
            """)
            result = cursor.fetchone()
            conn.close()
            self.assertIsNotNone(result, "ide_tooltip_table does not exist in the database")
        finally:
            os.remove(db_path)

    def test_corrupt_xlsx_file(self):
        # Create a temporary text file (not an xlsx)
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            tmp.write(b"This is not an xlsx file")
            xlsx_path = tmp.name

        # Create a temporary database file
        with tempfile.NamedTemporaryFile(delete=False) as tmp:
            db_path = tmp.name

        try:
            # Attempt to create TooltipDatabase with corrupt xlsx
            db = TooltipDatabase(db_path, xlsx_path)
            with self.assertRaises(Exception):
                db.read_xlsx()  # This method will need to be implemented
        finally:
            # Clean up temporary files
            os.remove(xlsx_path)
            os.remove(db_path)

    def test_invalid_xlsx_headers(self):
        # Create a temporary xlsx file with incorrect headers
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            xlsx_path = tmp.name

        # Create a workbook and add incorrect headers
        wb = Workbook()
        ws = wb.active
        ws.append(['wrong', 'headers', 'here'])  # Incorrect headers
        wb.save(xlsx_path)

        # Create a temporary database file
        with tempfile.NamedTemporaryFile(delete=False) as tmp:
            db_path = tmp.name

        try:
            # Attempt to create TooltipDatabase with invalid headers
            db = TooltipDatabase(db_path, xlsx_path)
            with self.assertRaises(Exception):
                db.read_xlsx()  # This method will need to be implemented
        finally:
            # Clean up temporary files
            os.remove(xlsx_path)
            os.remove(db_path)

    def test_empty_cells_in_row(self):
        # Create a temporary xlsx file with correct headers but empty cells
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            xlsx_path = tmp.name

        # Create a temporary database file
        with tempfile.NamedTemporaryFile(delete=False) as tmp:
            db_path = tmp.name

        try:
            # Test each required column being empty
            required_columns = ['tag', 'category', 'summary', 'buttonDescr1', 'buttonURI1']
            
            for empty_col in required_columns:
                with self.subTest(empty_column=empty_col):
                    # Create a workbook with correct headers
                    wb = Workbook()
                    ws = wb.active
                    
                    # Add header row
                    ws.append(TooltipDatabase.REQUIRED_HEADERS)
                    
                    # Add a row with all values filled except the current column
                    row = ['tag', 'category', 'summary', 'detail', 'button1', 'uri1', '', '', '', '']
                    # Set the current column to empty
                    col_index = TooltipDatabase.REQUIRED_HEADERS.index(empty_col)
                    row[col_index] = ''
                    
                    ws.append(row)
                    wb.save(xlsx_path)

                    # Attempt to create TooltipDatabase with empty cell
                    db = TooltipDatabase(db_path, xlsx_path)
                    with self.assertRaises(Exception):
                        db.read_xlsx()  # This method will need to be implemented
        finally:
            # Clean up temporary files
            os.remove(xlsx_path)
            os.remove(db_path)

    def test_read_xlsx_with_random_content(self):
        # Create a temporary xlsx file
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            xlsx_path = tmp.name

        # Create a database file with a fixed name for inspection
        db_path = "test_tooltips.db"
        
        # Remove old database file if it exists
        if os.path.exists(db_path):
            os.remove(db_path)

        try:
            # Create a workbook with correct headers
            wb = Workbook()
            ws = wb.active
            
            # Add header row
            ws.append(TooltipDatabase.REQUIRED_HEADERS)
            
            # Add 5 rows of random content
            import random
            import string
            
            def random_string(length=8):
                return ''.join(random.choices(string.ascii_letters + string.digits, k=length))
            
            def random_url():
                return f"http://localhost:6174/{random_string()}"
            
            # First row: all buttons present
            ws.append([
                random_string(),  # tag
                random_string(),  # category
                random_string(),  # summary
                random_string(),  # detail
                random_string(),  # buttonDescr1
                random_url(),     # buttonURI1
                random_string(),  # buttonDescr2
                random_url(),     # buttonURI2
                random_string(),  # buttonDescr3
                random_url(),     # buttonURI3
            ])
            
            # Second row: missing button3
            ws.append([
                random_string(),  # tag
                random_string(),  # category
                random_string(),  # summary
                random_string(),  # detail
                random_string(),  # buttonDescr1
                random_url(),     # buttonURI1
                random_string(),  # buttonDescr2
                random_url(),     # buttonURI2
                '',               # buttonDescr3 (empty)
                '',               # buttonURI3 (empty)
            ])
            
            # Third row: missing both button2 and button3
            ws.append([
                random_string(),  # tag
                random_string(),  # category
                random_string(),  # summary
                random_string(),  # detail
                random_string(),  # buttonDescr1
                random_url(),     # buttonURI1
                '',               # buttonDescr2 (empty)
                '',               # buttonURI2 (empty)
                '',               # buttonDescr3 (empty)
                '',               # buttonURI3 (empty)
            ])
            
            # Fourth and fifth rows: all buttons present
            for _ in range(2):
                ws.append([
                    random_string(),  # tag
                    random_string(),  # category
                    random_string(),  # summary
                    random_string(),  # detail
                    random_string(),  # buttonDescr1
                    random_url(),     # buttonURI1
                    random_string(),  # buttonDescr2
                    random_url(),     # buttonURI2
                    random_string(),  # buttonDescr3
                    random_url(),     # buttonURI3
                ])
            
            wb.save(xlsx_path)

            # Create database and read xlsx
            db = TooltipDatabase(db_path, xlsx_path)
            db.read_xlsx()

            # Verify the data was inserted correctly
            conn = sqlite3.connect(db_path)
            cursor = conn.cursor()
            cursor.execute("SELECT COUNT(*) FROM ide_tooltip_table")
            count = cursor.fetchone()[0]
            conn.close()

            self.assertEqual(count, 5, "Expected 5 rows in the database")

        finally:
            # Only clean up the xlsx file
            os.remove(xlsx_path)

    def test_tag_with_false_string(self):
        # Create a temporary xlsx file
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            xlsx_path = tmp.name

        # Create a database file with a fixed name for inspection
        db_path = "test_tooltips.db"
        
        # Remove old database file if it exists
        if os.path.exists(db_path):
            os.remove(db_path)

        try:
            # Create a workbook with correct headers
            wb = Workbook()
            ws = wb.active
            
            # Add header row
            ws.append(TooltipDatabase.REQUIRED_HEADERS)
            
            # Add a row with "FALSE" as the tag
            ws.append([
                "FALSE",           # tag
                "test_category",   # category
                "test_summary",    # summary
                "test_detail",     # detail
                "button1",         # buttonDescr1
                "http://localhost:6174/test1",  # buttonURI1
                "",               # buttonDescr2
                "",               # buttonURI2
                "",               # buttonDescr3
                "",               # buttonURI3
            ])
            
            wb.save(xlsx_path)

            # Create database and read xlsx
            db = TooltipDatabase(db_path, xlsx_path)
            db.read_xlsx()

            # Verify the data was inserted correctly
            conn = sqlite3.connect(db_path)
            cursor = conn.cursor()
            cursor.execute("SELECT tooltipTag FROM ide_tooltip_table")
            result = cursor.fetchone()
            conn.close()

            self.assertEqual(result[0], "FALSE", "Tag value 'FALSE' was not preserved correctly")

        finally:
            # Clean up temporary files
            os.remove(xlsx_path)

if __name__ == '__main__':
    unittest.main()
